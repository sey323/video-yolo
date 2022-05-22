import base64
import logging
import os
from datetime import datetime as dt

import cv2
import numpy as np
import openpyxl

logger = logging.getLogger(__name__)


class ImageUtil(object):
    @staticmethod
    def save_image_from_base64(image_base64: str, save_path: str) -> str:
        img_binary = base64.b64decode(image_base64)
        jpg = np.frombuffer(img_binary, dtype=np.uint8)

        # raw image <- jpg
        img = cv2.imdecode(jpg, cv2.IMREAD_COLOR)

        # 画像の保存
        os.makedirs(save_path, exist_ok=True)
        img_name = dt.now().strftime("%Y%m%d%H%M%S") + ".png"
        save_image_path = os.path.join(save_path, img_name)
        cv2.imwrite(save_image_path, img)

        return save_image_path


class Dumper(object):
    """動画の検出結果をHTMLとExelファイルに書き込みを実施するクラス。

    Args:
        object ([type]): [description]
    """

    def __init__(self, save_path: str, img_size: tuple = (160, 90)):
        self.dump_text = ""
        self.bootstrap_cdn = '<head>\n \
                                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.6/css/bootstrap.min.css" />\n \
                                <script src="https://ajax.googleapis.com/ajax/libs/jquery/1.11.1/jquery.min.js"></script>\n \
                                <script src="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.6/js/bootstrap.min.js"></script>\n \
                            </head>'
        self.begin = '<table class="table table-responsive">'
        self.end = "</table>"
        self.content = ""
        self.img_size = img_size
        # openpyxl
        self.work_book = openpyxl.Workbook()
        self.work_sheet = self.work_book.worksheets[0]
        self.work_sheet.column_dimensions["C"].width = img_size[0] * 0.125
        self.work_sheet.column_dimensions["D"].width = 30

        # 初期設定
        self.work_sheet.cell(row=1, column=1).value = "No."
        self.work_sheet.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(
            vertical="center", horizontal="center"
        )
        self.work_sheet.cell(row=1, column=2).value = "時間"
        self.work_sheet.cell(row=1, column=2).alignment = openpyxl.styles.Alignment(
            vertical="center", horizontal="center"
        )
        self.work_sheet.cell(row=1, column=3).value = "場面"
        self.work_sheet.cell(row=1, column=3).alignment = openpyxl.styles.Alignment(
            vertical="center", horizontal="center"
        )
        self.work_sheet.cell(row=1, column=4).value = "撮影対象物"
        self.work_sheet.cell(row=1, column=4).alignment = openpyxl.styles.Alignment(
            vertical="center", horizontal="center"
        )

        self.no = 1

        self.save_path = save_path
        self.save_images_path = os.path.join(self.save_path, "material")

        # フォルダを作成する
        os.makedirs(self.save_path, exist_ok=True)
        os.makedirs(os.path.join(self.save_images_path), exist_ok=True)

    def _save_image(self, image: np.ndarray, image_name: str):
        """指定したサイズに画像をリサイズし、Dumberように画像を保存する

        Args:
            image (np.ndarray): 保存する画像
            image_name (str): 画像ファイル名
        """
        save_img_path = os.path.join(self.save_images_path, "{}.jpg".format(image_name))
        # ExelやHTMLに保存するように画像を保存
        resized_image = cv2.resize(
            image, self.img_size, interpolation=cv2.INTER_AREA
        )  # 指定したサイズに画像を縮小
        cv2.imwrite(save_img_path, resized_image)

    def add_scene(self, frame: str, time: str, image: np.ndarray, param_path: str):
        """HTMLとExelにシーンと検出結果を追記する

        Args:
            frame (str): フレーム番号
            time (str): 該当シーンの動画の経過時間
            image (np.ndarray): HTMLやXMLに埋め込む画像
            param_path (str): 読み込むパラメータファイルのパス
        """
        self._save_image(image, frame)

        # HTMLに出力
        with open(param_path, "r", encoding="utf-8") as f:
            param_data = f.read()
        self.content += '<tr class="col-md-12">\n \
                            <td class="col-md-1">{0}</td>\n \
                            <td class="col-md-1">{1}</td>\n \
                            <td class="col-md-5">\n  \
                                <img class="gen-img" src="{2}">\n \
                            </td>\n \
                            <td class="col-md-4">{3}</td>\n \
                            <td class="col-md-2"></td>\n \
                        </tr>'.format(
            self.no,
            time,
            os.path.join("material", "{}.jpg".format(str(frame))),
            param_data,
        )

        # exelに出力
        exel_no = self.no + 1
        self.work_sheet.cell(row=exel_no, column=1).value = self.no
        self.work_sheet.cell(
            row=exel_no, column=1
        ).alignment = openpyxl.styles.Alignment(vertical="center")
        self.work_sheet.cell(row=exel_no, column=2).value = time
        self.work_sheet.cell(
            row=exel_no, column=2
        ).alignment = openpyxl.styles.Alignment(vertical="center")
        self.work_sheet.cell(row=exel_no, column=4).value = param_data.replace(
            "<br>", "\n"
        )
        self.work_sheet.cell(
            row=exel_no, column=4
        ).alignment = openpyxl.styles.Alignment(vertical="center", wrapText=True)

        # Exelに画像の入力
        self.work_sheet.row_dimensions[exel_no].height = self.img_size[1] * 0.78
        img = openpyxl.drawing.image.Image(
            os.path.join(self.save_images_path, "{}.jpg".format(str(frame)))
        )
        img.anchor = self.work_sheet.cell(row=exel_no, column=3)
        img.anchor = "C" + str(exel_no)
        self.work_sheet.add_image(img)

        self.no += 1

    def save_html(self):
        """作成したHTMLとxlsxファイルを保存"""
        html = self.begin + self.content + self.end
        with open(os.path.join(self.save_path, "index.html"), mode="w") as f:
            f.write(html)

        self.work_book.save(os.path.join(self.save_path, "result.xlsx"))
        logger.info(f"Html save complete!: {self.save_path}")
